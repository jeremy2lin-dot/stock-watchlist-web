from __future__ import annotations

import json
import os
import secrets
import tempfile
from io import BytesIO
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from watchlist_providers import (
    GoodinfoExperimentalProvider,
    MegaApiProvider,
    ProviderResult,
    TwseTpexProvider,
)


APP_DIR = Path(__file__).resolve().parent
DATA_PATH = Path(os.getenv("WATCHLIST_DATA_PATH", APP_DIR / "watchlist_data.json"))
DEFAULT_DATA_PATH = APP_DIR / "default_watchlist_data.json"
EXPORT_DIR = Path(os.getenv("WATCHLIST_EXPORT_DIR", DATA_PATH.parent / "exports"))
APP_USER = os.getenv("WATCHLIST_USER", "")
APP_PASSWORD = os.getenv("WATCHLIST_PASSWORD", "")

app = Flask(__name__)

EXPORT_COLUMNS = [
    ("watch_date", "\u89c0\u5bdf\u65e5\u671f"),
    ("ticker", "\u80a1\u7968\u4ee3\u865f"),
    ("name", "\u80a1\u7968\u540d\u7a31"),
    ("price", "\u73fe\u50f9"),
    ("planned_buy_price", "\u9810\u5b9a\u8cb7\u5165\u50f9\u4f4d"),
    ("ma5", "5MA"),
    ("ma10", "10MA"),
    ("ma20", "20MA"),
    ("ma50", "50MA"),
    ("entry", "\u640d\u5e73\u50f9"),
    ("stop_loss", "\u505c\u640d\u50f9"),
    ("take_profit", "\u505c\u5229\u50f9"),
    ("action", "Action"),
    ("trend", "Trend"),
    ("strategy", "Strategy"),
    ("strategy_status", "\u7b56\u7565\u72c0\u614b"),
    ("user_notes", "\u8a3b\u8a18"),
    ("last_update", "\u6700\u5f8c\u66f4\u65b0"),
    ("notes", "\u7cfb\u7d71\u5099\u8a3b"),
]


def require_auth(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not APP_USER and not APP_PASSWORD:
            return func(*args, **kwargs)

        auth = request.authorization
        user_ok = auth and secrets.compare_digest(auth.username or "", APP_USER)
        password_ok = auth and secrets.compare_digest(auth.password or "", APP_PASSWORD)
        if user_ok and password_ok:
            return func(*args, **kwargs)

        return Response(
            "Authentication required",
            401,
            {"WWW-Authenticate": 'Basic realm="Watchlist"'},
        )

    return wrapper


def normalize_stock_code(value: object) -> str:
    raw = str(value).strip()
    if not raw:
        return ""
    if raw.endswith(".0"):
        raw = raw[:-2]
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits:
        return raw
    return digits.zfill(4) if len(digits) < 4 else digits


def parse_float(value: object) -> float | None:
    if value in (None, "", "--", "---", "----", "X0.00"):
        return None
    text = str(value).replace(",", "").replace("X", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_row(row: dict) -> dict:
    row.setdefault("watch_date", "")
    row["ticker"] = normalize_stock_code(row.get("ticker", ""))
    row.setdefault("name", "")
    row.setdefault("price", "")
    row.setdefault("planned_buy_price", "")
    row.setdefault("ma5", "")
    row.setdefault("ma10", "")
    row.setdefault("ma20", "")
    row.setdefault("ma50", "")
    row.setdefault("entry", "")
    row.setdefault("stop_loss", "")
    row.setdefault("take_profit", "")
    row.setdefault("action", "")
    row.setdefault("trend", "")
    row.setdefault("strategy", "")
    row.setdefault("strategy_status", "")
    row.setdefault("user_notes", "")
    row.setdefault("last_update", "")
    row.setdefault("notes", "")
    return row


def normalize_data(data: dict) -> dict:
    data.setdefault("provider", "twse_tpex")
    data["rows"] = [normalize_row(row) for row in data.get("rows", [])]
    return data


def load_data() -> dict:
    if DATA_PATH.exists():
        return normalize_data(json.loads(DATA_PATH.read_text(encoding="utf-8")))
    if DEFAULT_DATA_PATH.exists():
        return normalize_data(json.loads(DEFAULT_DATA_PATH.read_text(encoding="utf-8")))
    return {"rows": [], "provider": "twse_tpex"}


def save_data(data: dict) -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = DATA_PATH.with_suffix(DATA_PATH.suffix + ".tmp")
    tmp_path.write_text(json.dumps(normalize_data(data), ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, DATA_PATH)


def import_rows_from_excel(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, keep_vba=True, data_only=False)
    ws = wb.worksheets[0]
    rows: list[dict] = []
    row_num = 5
    while True:
        ticker = ws[f"B{row_num}"].value
        watch_date = ws[f"A{row_num}"].value
        entry = ws[f"I{row_num}"].value
        if ticker in (None, "") and watch_date in (None, "") and entry in (None, ""):
            break

        def cell(col: str) -> str:
            value = ws[f"{col}{row_num}"].value
            return "" if value is None else str(value).strip()

        rows.append(
            normalize_row(
                {
                    "watch_date": cell("A"),
                    "ticker": cell("B"),
                    "name": cell("C"),
                    "price": cell("D"),
                    "planned_buy_price": "",
                    "ma5": cell("E"),
                    "ma10": cell("F"),
                    "ma20": cell("G"),
                    "ma50": cell("H"),
                    "entry": cell("I"),
                    "stop_loss": cell("J"),
                    "take_profit": cell("K"),
                    "action": cell("L"),
                    "trend": cell("M"),
                    "strategy": "",
                    "strategy_status": cell("N"),
                    "last_update": cell("O"),
                    "user_notes": "",
                    "notes": cell("P"),
                }
            )
        )
        row_num += 1
    return rows


def export_rows_to_excel(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="174D7A")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row in enumerate(data.get("rows", []), start=2):
        normalize_row(row)
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_alignment

    widths = {
        "A": 14,
        "B": 12,
        "C": 16,
        "D": 12,
        "E": 18,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 14,
        "N": 14,
        "O": 42,
        "P": 18,
        "Q": 42,
        "R": 20,
        "S": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def apply_row_strategy_fields(row: dict, result: ProviderResult) -> None:
    stop_loss = parse_float(row.get("stop_loss"))
    take_profit = parse_float(row.get("take_profit"))
    planned_buy_price = parse_float(row.get("planned_buy_price"))
    price = parse_float(row.get("price"))

    if price is None:
        action = "\u89c0\u5bdf"
    elif take_profit is not None and price >= take_profit:
        action = "\u505c\u5229"
    elif stop_loss is not None and price <= stop_loss:
        action = "\u505c\u640d"
    elif planned_buy_price is not None and price <= planned_buy_price:
        action = "\u53ef\u8cb7\u5165"
    else:
        action = "\u7e8c\u62b1/\u89c0\u5bdf"

    row["action"] = action
    if row.get("ma5") in ("", None) or row.get("ma10") in ("", None) or row.get("ma20") in ("", None) or row.get("ma50") in ("", None):
        row["trend"] = "\u8cc7\u6599\u4e0d\u8db3"
    else:
        row["trend"] = result.trend

    if action == "\u505c\u5229":
        row["strategy_status"] = "\u9054\u5230\u505c\u5229\u50f9"
    elif action == "\u505c\u640d":
        row["strategy_status"] = "\u8dcc\u7834\u505c\u640d\u50f9"
    elif action == "\u53ef\u8cb7\u5165":
        row["strategy_status"] = "\u4f4e\u65bc\u9810\u5b9a\u8cb7\u5165\u50f9"
    else:
        row["strategy_status"] = "\u6301\u7e8c\u8ffd\u8e64"


def provider_by_name(name: str):
    if name == "goodinfo_exp":
        return GoodinfoExperimentalProvider()
    if name == "mega_api":
        return MegaApiProvider()
    return TwseTpexProvider()


@app.get("/")
@require_auth
def index():
    return render_template("index.html")


@app.get("/health")
def health():
    return jsonify({"ok": True})


@app.get("/api/watchlist")
@require_auth
def api_watchlist():
    return jsonify(load_data())


@app.get("/api/providers")
@require_auth
def api_providers():
    return jsonify(
        {
            "default": "twse_tpex",
            "providers": [
                {"id": "twse_tpex", "name": "TWSE/TPEX (\u516c\u958b\u8cc7\u6599)"},
                {"id": "goodinfo_exp", "name": "Goodinfo \u5be6\u9a57\u6a21\u5f0f\uff08\u975e\u6b63\u5f0f\uff09"},
                {"id": "mega_api", "name": "\u5146\u8c50 API (\u672c\u6a5f Speedy \u884c\u60c5)"},
            ],
        }
    )


@app.get("/api/diagnostics/<code>")
@require_auth
def api_diagnostics(code: str):
    normalized = normalize_stock_code(code)
    provider = TwseTpexProvider()
    result = provider.fetch(normalized, {})
    return jsonify(
        {
            "ticker": normalized,
            "name": result.name,
            "price": result.price,
            "ma5": result.ma5,
            "ma10": result.ma10,
            "ma20": result.ma20,
            "ma50": result.ma50,
            "trend": result.trend,
            "note": result.note,
        }
    )


@app.post("/api/provider")
@require_auth
def api_provider():
    payload = request.get_json(force=True)
    provider = str(payload.get("provider", "twse_tpex")).strip()
    data = load_data()
    data["provider"] = provider
    save_data(data)
    return jsonify({"ok": True, "provider": provider})


@app.post("/api/watchlist")
@require_auth
def api_save_watchlist():
    payload = request.get_json(force=True)
    rows = payload.get("rows", [])
    data = load_data()
    data["rows"] = rows
    save_data(data)
    return jsonify({"ok": True})


@app.post("/api/update")
@require_auth
def api_update():
    data = load_data()
    rows: list[dict] = data.get("rows", [])
    payload = request.get_json(silent=True) or {}
    provider_name = str(payload.get("provider") or data.get("provider", "twse_tpex"))
    data["provider"] = provider_name
    provider = provider_by_name(provider_name)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    success_count = 0
    failed_count = 0
    first_failure = ""

    for row in rows:
        normalize_row(row)
        code = normalize_stock_code(row.get("ticker", ""))
        row["ticker"] = code
        if not code:
            continue

        result = provider.fetch(code, row)
        row["name"] = result.name or row.get("name", "")
        row["price"] = result.price if result.price is not None else row.get("price", "")
        row["ma5"] = result.ma5 if result.ma5 is not None else ""
        row["ma10"] = result.ma10 if result.ma10 is not None else ""
        row["ma20"] = result.ma20 if result.ma20 is not None else ""
        row["ma50"] = result.ma50 if result.ma50 is not None else ""
        row["notes"] = result.note
        row["last_update"] = now
        apply_row_strategy_fields(row, result)
        if row["ma5"] != "" and row["ma10"] != "" and row["ma20"] != "" and row["ma50"] != "":
            success_count += 1
        else:
            failed_count += 1
            if not first_failure:
                first_failure = f"{code}: {result.note}"

    data["rows"] = rows
    save_data(data)
    return jsonify(
        {
            "ok": True,
            "rows": rows,
            "provider": provider_name,
            "success_count": success_count,
            "failed_count": failed_count,
            "first_failure": first_failure,
        }
    )


@app.post("/api/import_excel")
@require_auth
def api_import_excel():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "missing_file"}), 400
    up = request.files["file"]
    if not up.filename:
        return jsonify({"ok": False, "error": "empty_filename"}), 400

    suffix = Path(up.filename).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        return jsonify({"ok": False, "error": "invalid_extension"}), 400

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            up.save(tmp.name)
            tmp_path = Path(tmp.name)
        rows = import_rows_from_excel(tmp_path)
        data = load_data()
        data["rows"] = rows
        save_data(data)
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


@app.get("/api/export_excel")
@require_auth
def api_export_excel():
    output = export_rows_to_excel(load_data())
    filename = f"stock_watchlist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    saved_path = EXPORT_DIR / filename
    saved_path.write_bytes(output.getvalue())
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host="127.0.0.1", port=5050, debug=debug)
